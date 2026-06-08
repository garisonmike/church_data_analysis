#!/usr/bin/env python3
"""
Church Data Analysis CLI.

This is the terminal companion to the Flutter app's analytics screens. It reads
weekly church records from CSV/XLSX files, normalizes app-compatible column names,
derives the same core metrics used by the app, and exports selectable graphs.

Quick examples (run from the terminalVersion/ folder):
    python data.py --input data/netFinalData.csv --graphs all --pdf
    python data.py --input data --group attendance --export-clean
    python data.py --input weekly.xlsx --graphs total_attendance_trend,income_distribution
    python data.py --list-graphs

Import template columns (Flutter app):
    week_start_date, men, women, youth, children, sunday_home_church,
    tithe, offerings, emergency_collection, planned_collection,
    baptisms, holy_communion

Extra columns supported by this CLI (not in the app template yet):
    sabbath_school_attendance, visitors_count, mission_offering,
    local_church_budget, total_attendance, total_income,
    id, church_id, created_by_admin_id, created_at, updated_at,
    board_business_meeting_attendance, board_business_meeting_expected,
    holy_communion_expected, ambassadors_attendance, adult_attendance

Optional metadata columns preserved when present:
    granularity, source, source_file, source_sheet, source_table

Notes:
    - CSV files are assumed to use the first row as headers.
    - XLSX/XLS files can have title rows; the script scans early rows for the
      most likely header row on each sheet.
    - PNGs are saved per graph. With --pdf, the same figures are also written
      into one multi-page PDF.
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Iterable, Optional, Sequence

os.environ.setdefault("MPLCONFIGDIR", "/tmp/matplotlib")

import matplotlib
from matplotlib.axes import Axes
from matplotlib.figure import Figure

if not os.environ.get("DISPLAY"):
    matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.backends.backend_pdf import PdfPages

SUPPORTED_SUFFIXES = {".csv", ".xlsx", ".xls", ".xslx"}
SCRIPT_DIR = Path(__file__).resolve().parent

# ---------------------------------------------------------------------------
# Church brand palette
# ---------------------------------------------------------------------------
C_NAVY   = "#1B3A6B"
C_GOLD   = "#C8972B"
C_SKY    = "#4A90D9"
C_GREEN  = "#2E8B57"
C_ROSE   = "#C0395A"
C_AMBER  = "#E07B30"
C_SLATE  = "#5C6E8A"
C_LIGHT  = "#EEF3FA"
C_WHITE  = "#FFFFFF"
C_GREY   = "#D0D7E2"

CHURCH_PALETTE = [C_NAVY, C_GOLD, C_SKY, C_GREEN, C_ROSE, C_AMBER, C_SLATE,
                  "#7B5EA7", "#2CA4A4", "#A44A3F", "#7A9E3B", "#4A7A7A"]

WATERMARK_TEXT = "Kisii Central SDA Church  ·  Q1 2026"

IMPORT_TEMPLATE_COLUMNS = [
    "week_start_date",
    "men",
    "women",
    "youth",
    "children",
    "sunday_home_church",
    "tithe",
    "offerings",
    "emergency_collection",
    "planned_collection",
    "baptisms",
    "holy_communion",
]

APP_WEEKLY_COLUMNS = [
    "id",
    "church_id",
    "created_by_admin_id",
    "week_start_date",
    "men",
    "women",
    "youth",
    "children",
    "sunday_home_church",
    "total_attendance",
    "tithe",
    "offerings",
    "emergency_collection",
    "planned_collection",
    "mission_offering",
    "local_church_budget",
    "total_income",
    "baptisms",
    "holy_communion",
    "holy_communion_expected",
    "sabbath_school_attendance",
    "visitors_count",
    "board_business_meeting_attendance",
    "board_business_meeting_expected",
    "ambassadors_attendance",
    "adult_attendance",
    "created_at",
    "updated_at",
]

OPTIONAL_METADATA_COLUMNS = [
    "granularity",
    "source",
    "source_file",
    "source_sheet",
    "source_table",
]

IGNORED_COLUMNS = {
    "col",
    "col_2",
    "col_3",
    "col_4",
    "col_5",
    "25",
    "112",
    "126",
    "148",
}

ATTENDANCE_PARTS = ["men", "women", "youth", "children", "sunday_home_church"]
SABBATH_ATTENDANCE_PARTS = ["men", "women", "youth", "children"]
CORE_INCOME_PARTS = ["tithe", "offerings", "emergency_collection", "planned_collection"]
OPTIONAL_INCOME_PARTS = ["mission_offering", "local_church_budget"]
INCOME_PARTS = CORE_INCOME_PARTS + OPTIONAL_INCOME_PARTS

NUMERIC_COLUMNS = {
    "id",
    "church_id",
    "created_by_admin_id",
    "men",
    "women",
    "youth",
    "children",
    "sunday_home_church",
    "total_attendance",
    "tithe",
    "offerings",
    "emergency_collection",
    "planned_collection",
    "mission_offering",
    "local_church_budget",
    "total_income",
    "baptisms",
    "holy_communion",
    "holy_communion_expected",
    "sabbath_school_attendance",
    "visitors_count",
    "board_business_meeting_attendance",
    "board_business_meeting_expected",
    "ambassadors_attendance",
    "adult_attendance",
}

HEADER_KEYWORDS = {
    "date",
    "week",
    "saturday",
    "men",
    "women",
    "youth",
    "children",
    "attendance",
    "home",
    "church",
    "tithe",
    "offering",
    "offerings",
    "emergency",
    "planned",
    "communion",
    "baptism",
    "board",
    "business",
    "meeting",
    "ambassador",
    "ambassadors",
}

ALIASES: dict[str, list[str]] = {
    "id": ["id", "record_id"],
    "church_id": ["church_id", "churchid"],
    "created_by_admin_id": ["created_by_admin_id", "createdbyadminid", "admin_id"],
    "week_start_date": [
        "week_start_date",
        "weekstartdate",
        "week_start",
        "week_starting",
        "week_date",
        "date",
        "saturday",
        "saturday_date",
        "service_date",
    ],
    "men": ["men", "male", "males", "men_attendance"],
    "women": ["women", "female", "females", "women_attendance"],
    "youth": ["youth", "youths", "youth_attendance", "teenagers"],
    "children": [
        "children",
        "child",
        "kids",
        "children_attendance",
        "total_all_children_classes",
        "total_in_all_children_classes",
    ],
    "sunday_home_church": [
        "sunday_home_church",
        "sundayhomechurch",
        "home_church",
        "homechurch",
        "home_ch",
        "sunday_home",
        "attendance_at_the_sunday_home_church_service_required_integer",
    ],
    "total_attendance": ["total_attendance", "totalattendance", "total_attend"],
    "tithe": ["tithe", "tithes", "tithing", "tithe_kes", "tithe_ksh"],
    "offerings": ["offerings", "offering", "offertory", "offerings_kes", "offering_kes"],
    "emergency_collection": [
        "emergency_collection",
        "emergencycollection",
        "emergency",
        "emergency_collection_kes",
    ],
    "planned_collection": [
        "planned_collection",
        "plannedcollection",
        "planned",
        "planned_collection_kes",
    ],
    "mission_offering": ["mission_offering", "missionoffering", "mission"],
    "local_church_budget": [
        "local_church_budget",
        "localchurchbudget",
        "church_budget",
        "local_budget",
    ],
    "total_income": ["total_income", "totalincome", "total_income_kes"],
    "baptisms": ["baptisms", "baptism", "baptised", "baptized"],
    "holy_communion": [
        "holy_communion",
        "holycommunion",
        "communion",
        "holy_comm",
        "hc",
        "hc_attendance",
        "quarterly_holy_communion_attendance",
    ],
    "holy_communion_expected": ["holy_communion_expected", "expected_at_hc"],
    "sabbath_school_attendance": [
        "sabbath_school_attendance",
        "sabbathschoolattendance",
        "sabbath_school",
        "sabbathschool",
        "kisii_central_sda_church_sabbath_school_attendance",
    ],
    "visitors_count": ["visitors_count", "visitorscount", "visitors", "visitor_count"],
    "board_business_meeting_attendance": [
        "board_business_meeting_attendance",
        "board_meeting",
        "board_meeting_attendance",
        "board_attended",
        "and_business_meeting_attendance",
        "business_mtg",
        "business_meeting_attendance",
        "kisii_central_sda_church_board_business_meeting_attendance_2026",
    ],
    "board_business_meeting_expected": [
        "board_business_meeting_expected",
        "board_meeting_expected",
        "board_expected",
    ],
    "ambassadors_attendance": ["ambassadors_attendance", "ambassadors"],
    "adult_attendance": ["adult_attendance", "adults"],
    "created_at": ["created_at", "createdat"],
    "updated_at": ["updated_at", "updatedat"],
}

ALIAS_LOOKUP = {alias: canonical for canonical, aliases in ALIASES.items() for alias in aliases}


def log(message: str) -> None:
    print(message)


def normalize_name(name: object) -> str:
    """Return a stable snake_case-ish key for matching loose spreadsheet headers."""
    text = str(name).strip().replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s*\([^)]*\)", "", text)
    text = text.encode("ascii", "ignore").decode()
    text = text.lower().replace("%", "percent")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return re.sub(r"_+", "_", text).strip("_")


def canonical_column_name(name: object) -> str:
    normalized = normalize_name(name)
    return ALIAS_LOOKUP.get(normalized, normalized)


def detect_header_row(raw: pd.DataFrame, scan_rows: int = 30) -> Optional[int]:
    """Find a likely header row in a raw Excel sheet that may contain title rows."""
    best_idx: Optional[int] = None
    best_score = 0
    for idx in range(min(len(raw), scan_rows)):
        tokens: set[str] = set()
        for value in raw.iloc[idx].tolist():
            if value is None:
                continue
            text = str(value).strip()
            if not text or text.lower() == "nan":
                continue
            tokens.update(re.split(r"[^a-zA-Z0-9]+", text.lower()))
        score = sum(1 for token in HEADER_KEYWORDS if token in tokens)
        if score >= 2 and score > best_score:
            best_idx = idx
            best_score = score
    return best_idx


def combine_duplicate_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Merge duplicate canonical columns, preferring the first non-empty value."""
    result = pd.DataFrame(index=df.index)
    for position, column in enumerate(df.columns):
        series = df.iloc[:, position]
        if column in result.columns:
            result[column] = result[column].combine_first(series)
        else:
            result[column] = series
    return result


def canonicalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [canonical_column_name(column) for column in df.columns]
    return combine_duplicate_columns(df)


def clean_number(value: object) -> object:
    if value is None:
        return np.nan
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none", "null", "-"}:
        return np.nan
    text = text.replace(",", "")
    text = re.sub(r"^[A-Z]{2,4}\s+", "", text)
    return text


def coerce_numeric_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for column in NUMERIC_COLUMNS.intersection(df.columns):
        df[column] = pd.to_numeric(df[column].map(clean_number), errors="coerce")
    return df


def parse_date_series(series: pd.Series, force_year: Optional[int] = None) -> pd.Series:
    """Parse ISO dates, day-first dates, and Excel serial dates."""
    if pd.api.types.is_numeric_dtype(series):
        numeric = series.dropna()
        if not numeric.empty and numeric.median() > 10000:
            parsed = pd.to_datetime(series, unit="D", origin="1899-12-30", errors="coerce")
        else:
            parsed = pd.to_datetime(series, errors="coerce")
    else:
        raw = series.astype(str).str.strip()
        dayfirst_pattern = r"^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}$"
        dayfirst = raw.str.match(dayfirst_pattern, na=False).mean() > 0.5
        parsed = pd.to_datetime(raw, errors="coerce", dayfirst=dayfirst)

    parsed = parsed.dt.tz_localize(None)
    if force_year is not None:
        parsed = parsed.apply(lambda value: value.replace(year=force_year) if pd.notna(value) else value)
    return parsed


def safe_divide(numerator: pd.Series, denominator: pd.Series) -> pd.Series:
    denominator = denominator.replace(0, np.nan)
    return numerator / denominator


def ensure_numeric_column(df: pd.DataFrame, column: str) -> pd.Series:
    if column not in df.columns:
        df[column] = np.nan
    return pd.to_numeric(df[column], errors="coerce")


def sum_existing(df: pd.DataFrame, columns: Sequence[str]) -> pd.Series:
    available = [column for column in columns if column in df.columns]
    if not available:
        return pd.Series(np.nan, index=df.index, dtype="float64")
    return df[available].apply(pd.to_numeric, errors="coerce").sum(axis=1, min_count=1)


def looks_like_sabbath_total(existing: pd.Series, sabbath_total: pd.Series, app_total: pd.Series) -> bool:
    """Detect legacy totals that excluded Sunday Home Church attendance."""
    comparable = pd.DataFrame(
        {"existing": existing, "sabbath": sabbath_total, "app": app_total}
    ).dropna()
    if comparable.empty:
        return False
    sabbath_matches = np.isclose(comparable["existing"], comparable["sabbath"], rtol=0.01, atol=1).mean()
    app_matches = np.isclose(comparable["existing"], comparable["app"], rtol=0.01, atol=1).mean()
    return sabbath_matches > 0.6 and sabbath_matches > app_matches


def derive_metrics(df: pd.DataFrame) -> pd.DataFrame:
    """Add app-style totals plus analysis-only helper columns."""
    df = df.copy()

    for column in APP_WEEKLY_COLUMNS:
        if column not in df.columns:
            df[column] = np.nan

    for column in NUMERIC_COLUMNS.intersection(df.columns):
        df[column] = pd.to_numeric(df[column], errors="coerce")

    df["sabbath_attendance"] = sum_existing(df, SABBATH_ATTENDANCE_PARTS)
    computed_total_attendance = sum_existing(df, ATTENDANCE_PARTS)

    if "total_attendance" in df.columns and df["total_attendance"].notna().any():
        df["source_total_attendance"] = df["total_attendance"]
        if looks_like_sabbath_total(df["total_attendance"], df["sabbath_attendance"], computed_total_attendance):
            df["total_attendance"] = computed_total_attendance.combine_first(df["total_attendance"])
        else:
            df["total_attendance"] = df["total_attendance"].combine_first(computed_total_attendance)
    else:
        df["total_attendance"] = computed_total_attendance

    df["total_with_home_church"] = computed_total_attendance
    if "adult_attendance" in df.columns:
        existing_adult = pd.to_numeric(df["adult_attendance"], errors="coerce")
    else:
        existing_adult = pd.Series(np.nan, index=df.index, dtype="float64")
    derived_adult = ensure_numeric_column(df, "men").fillna(0) + ensure_numeric_column(df, "women").fillna(0)
    df["adult_attendance"] = existing_adult.combine_first(derived_adult)
    df["young_attendance"] = ensure_numeric_column(df, "youth").fillna(0) + ensure_numeric_column(df, "children").fillna(0)

    df["core_income"] = sum_existing(df, CORE_INCOME_PARTS)
    df["regular_income"] = ensure_numeric_column(df, "tithe").fillna(0) + ensure_numeric_column(df, "offerings").fillna(0)
    df["special_collections"] = (
        ensure_numeric_column(df, "emergency_collection").fillna(0)
        + ensure_numeric_column(df, "planned_collection").fillna(0)
    )

    computed_total_income = sum_existing(df, INCOME_PARTS)
    if "total_income" in df.columns and df["total_income"].notna().any():
        df["source_total_income"] = df["total_income"]
        # The app derives totalIncome from its component fields, including the
        # newer optional giving streams when present.
        df["total_income"] = computed_total_income.combine_first(df["total_income"])
    else:
        df["total_income"] = computed_total_income

    df["income_per_attendee"] = safe_divide(df["total_income"], df["total_attendance"])
    df["tithe_per_attendee"] = safe_divide(df["tithe"], df["total_attendance"])
    df["offerings_per_attendee"] = safe_divide(df["offerings"], df["total_attendance"])
    df["regular_income_per_adult"] = safe_divide(df["regular_income"], df["adult_attendance"])

    pct_denominator = df["sabbath_attendance"].replace(0, np.nan)
    for source, target in [
        ("men", "men_pct"),
        ("women", "women_pct"),
        ("youth", "youth_pct"),
        ("children", "children_pct"),
    ]:
        df[target] = safe_divide(df[source], pct_denominator) * 100

    if "sunday_home_church" in df.columns:
        df["home_church_pct"] = safe_divide(df["sunday_home_church"], df["total_attendance"]) * 100

    df = df.sort_values("week_start_date", na_position="last").reset_index(drop=True)
    df["attendance_growth"] = df["total_attendance"].pct_change(fill_method=None) * 100
    df["income_growth"] = df["total_income"].pct_change(fill_method=None) * 100
    df["tithe_growth"] = df["tithe"].pct_change(fill_method=None) * 100
    df["men_women_ratio"] = safe_divide(df["men"], df["women"])
    df["adult_young_ratio"] = safe_divide(df["adult_attendance"], df["young_attendance"])
    df["tithe_offerings_ratio"] = safe_divide(df["tithe"], df["offerings"])

    return df


def prepare_dataframe(df: pd.DataFrame, source_file: str, source_sheet: Optional[str], force_year: Optional[int]) -> pd.DataFrame:
    df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
    if df.empty:
        return df

    df = canonicalize_columns(df)
    if IGNORED_COLUMNS.intersection(df.columns):
        df = df.drop(columns=[column for column in IGNORED_COLUMNS if column in df.columns])
    df = coerce_numeric_columns(df)

    if "week_start_date" in df.columns:
        df["week_start_date"] = parse_date_series(df["week_start_date"], force_year)
    else:
        df["week_start_date"] = pd.NaT

    if "source_file" in df.columns:
        df["source_file"] = df["source_file"].combine_first(pd.Series(source_file, index=df.index))
    else:
        df["source_file"] = source_file
    if source_sheet is not None:
        if "source_sheet" in df.columns:
            df["source_sheet"] = df["source_sheet"].combine_first(pd.Series(source_sheet, index=df.index))
        else:
            df["source_sheet"] = source_sheet

    return df.dropna(axis=0, how="all")


def read_csv_file(path: Path, force_year: Optional[int]) -> list[pd.DataFrame]:
    df = pd.read_csv(path)
    return [prepare_dataframe(df, path.name, None, force_year)]


def read_excel_file(path: Path, force_year: Optional[int]) -> list[pd.DataFrame]:
    frames: list[pd.DataFrame] = []
    xls = pd.ExcelFile(path)
    for sheet in xls.sheet_names:
        raw = pd.read_excel(xls, sheet_name=sheet, header=None)
        raw = raw.dropna(axis=0, how="all").dropna(axis=1, how="all")
        if raw.empty:
            continue

        header_idx = detect_header_row(raw)
        if header_idx is None:
            df = pd.read_excel(xls, sheet_name=sheet)
        else:
            df = raw.iloc[header_idx + 1 :].copy()
            df.columns = raw.iloc[header_idx].tolist()

        prepared = prepare_dataframe(df, path.name, str(sheet), force_year)
        if not prepared.empty:
            frames.append(prepared)
    return frames


def resolve_input_paths(inputs: Sequence[str]) -> list[Path]:
    paths: list[Path] = []
    for value in inputs:
        path = Path(value).expanduser()
        if not path.is_absolute() and not path.exists():
            path = SCRIPT_DIR / path
        if path.is_dir():
            for suffix in SUPPORTED_SUFFIXES:
                paths.extend(sorted(path.glob(f"*{suffix}")))
        elif path.exists():
            paths.append(path)
        else:
            log(f"Warning: input path does not exist: {path}")

    unique_paths: list[Path] = []
    seen: set[Path] = set()
    for path in paths:
        resolved = path.resolve()
        if resolved not in seen and path.suffix.lower() in SUPPORTED_SUFFIXES:
            unique_paths.append(path)
            seen.add(resolved)
    return unique_paths


def load_data(paths: Sequence[Path], force_year: Optional[int]) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    for path in paths:
        suffix = path.suffix.lower()
        try:
            if suffix == ".csv":
                frames.extend(read_csv_file(path, force_year))
            elif suffix in {".xlsx", ".xls", ".xslx"}:
                frames.extend(read_excel_file(path, force_year))
        except Exception as exc:
            log(f"Warning: could not read {path}: {exc}")

    if not frames:
        return pd.DataFrame()

    combined = pd.concat(frames, ignore_index=True, sort=False)
    combined = combine_duplicate_columns(combined)
    combined = derive_metrics(combined)

    metric_columns = [
        "men",
        "women",
        "youth",
        "children",
        "sunday_home_church",
        "tithe",
        "offerings",
        "total_attendance",
        "total_income",
        "baptisms",
        "holy_communion",
        "sabbath_school_attendance",
    ]
    has_metric = combined[[c for c in metric_columns if c in combined.columns]].notna().any(axis=1)
    return combined[has_metric].reset_index(drop=True)


def has_data(df: pd.DataFrame, columns: Iterable[str]) -> bool:
    for column in columns:
        if column not in df.columns or df[column].dropna().empty:
            return False
    return True


def date_labels(df: pd.DataFrame) -> list[str]:
    if "week_start_date" not in df.columns:
        return [str(i + 1) for i in range(len(df))]
    dates = pd.to_datetime(df["week_start_date"], errors="coerce")
    labels = dates.dt.strftime("%Y-%m-%d")
    return [label if isinstance(label, str) else f"Row {i + 1}" for i, label in enumerate(labels)]


def configure_axes(ax: Axes, title: str, ylabel: Optional[str] = None) -> None:
    """Legacy plain style – kept so existing builders still work unchanged."""
    ax.set_title(title)
    if ylabel:
        ax.set_ylabel(ylabel)
    ax.grid(True, alpha=0.25)


def church_style(fig: Figure, ax: Axes, title: str) -> None:
    """Apply the church brand style to a figure / axes pair."""
    fig.patch.set_facecolor(C_WHITE)
    ax.set_facecolor(C_LIGHT)
    ax.set_title(title, fontsize=13, fontweight="bold", color=C_NAVY, pad=12)
    ax.tick_params(labelsize=9, colors="#3A3A3A")
    for spine in ax.spines.values():
        spine.set_edgecolor(C_GREY)
        spine.set_linewidth(0.8)
    ax.yaxis.grid(True, color=C_GREY, linewidth=0.6, linestyle="--", alpha=0.8)
    ax.set_axisbelow(True)


def add_watermark(fig: Figure) -> None:
    fig.text(0.99, 0.01, WATERMARK_TEXT,
             ha="right", va="bottom", fontsize=7.5, color="#AAAAAA", style="italic")


def bar_value_labels(ax: Axes, bars: list, color: str = C_NAVY, fmt: str = "{:.0f}") -> None:
    """Place value labels above each bar."""
    for bar in bars:
        h = bar.get_height()
        if pd.notna(h) and h > 0:
            ax.text(bar.get_x() + bar.get_width() / 2, h + ax.get_ylim()[1] * 0.01,
                    fmt.format(h), ha="center", va="bottom",
                    fontsize=8, fontweight="bold", color=color)


def make_church_fig(figsize: tuple = (12, 5)) -> tuple:
    """Return a (fig, ax) pair with the church background already applied."""
    fig, ax = plt.subplots(figsize=figsize)
    fig.patch.set_facecolor(C_WHITE)
    ax.set_facecolor(C_LIGHT)
    ax.yaxis.grid(True, color=C_GREY, linewidth=0.6, linestyle="--", alpha=0.8)
    ax.set_axisbelow(True)
    for spine in ax.spines.values():
        spine.set_edgecolor(C_GREY)
        spine.set_linewidth(0.8)
    return fig, ax


def set_categorical_x(ax: Axes, labels: Sequence[str]) -> np.ndarray:
    x = np.arange(len(labels))
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=25, ha="right")
    return x


def numeric_series(df: pd.DataFrame, column: str) -> pd.Series:
    return pd.to_numeric(df[column], errors="coerce")


def clean_positive_pairs(df: pd.DataFrame, x_col: str, y_col: str) -> pd.DataFrame:
    clean = df[[x_col, y_col]].apply(pd.to_numeric, errors="coerce").dropna()
    return clean[np.isfinite(clean[x_col]) & np.isfinite(clean[y_col])]


def add_mean_line(ax: Axes, series: pd.Series, label: str = "Average") -> None:
    mean = series.dropna().mean()
    if pd.notna(mean):
        ax.axhline(mean, linestyle="--", linewidth=1, color="#555555", label=label)


# ---------------------------------------------------------------------------
# ── NEW PRESENTATION GRAPHS ─────────────────────────────────────────────────
# These builders read the raw XLSX source files from a `data/` sub-folder
# next to this script when the aggregated CSV lacks that granularity
# (Sabbath School groups, Home Church breakdown, Board/Business meeting per
#  home-church). For the weekly-series graphs they use the normalised df.
# ---------------------------------------------------------------------------

def _data_dir() -> Path:
    """Return the data/ folder next to this script, or cwd/data as fallback."""
    for candidate in [SCRIPT_DIR / "data", Path.cwd() / "data"]:
        if candidate.is_dir():
            return candidate
    return SCRIPT_DIR  # last resort: same folder


def _find_xlsx(pattern: str) -> Optional[Path]:
    """Case-insensitive glob for an XLSX file matching *pattern* in data/."""
    d = _data_dir()
    pattern_lower = pattern.lower()
    for p in d.glob("*.xlsx"):
        if pattern_lower in p.name.lower():
            return p
    return None


# ── 1. Sabbath School group bar chart ───────────────────────────────────────

def plot_sabbath_school_groups(df: pd.DataFrame) -> Optional[Figure]:
    """
    Grouped bar chart: each Sabbath School study group × up to 3 recorded
    sessions.  Reads SABBATH SCHOOL DATA*.xlsx from data/.  Falls back to
    showing a single-bar chart from the weekly total column if the XLSX is
    absent.
    """
    path = _find_xlsx("sabbath school")
    if path is not None:
        try:
            return _sabbath_school_from_xlsx(path)
        except Exception as exc:
            log(f"Note: could not parse Sabbath School XLSX ({exc}), falling back to summary.")

    # Fallback: use sabbath_school_attendance column
    if not has_data(df, ["sabbath_school_attendance"]):
        return None
    labels = date_labels(df)
    fig, ax = make_church_fig()
    bars = ax.bar(labels, df["sabbath_school_attendance"].fillna(0), color=C_NAVY, zorder=3)
    bar_value_labels(ax, bars)
    church_style(fig, ax, "Sabbath School Attendance by Week")
    ax.set_ylabel("Attendance", fontsize=10)
    ax.tick_params(axis="x", rotation=25)
    add_watermark(fig)
    return fig


def _sabbath_school_from_xlsx(path: Path) -> Optional[Figure]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    if ws is None:
        return None
    rows = list(ws.iter_rows(values_only=True))

    # Locate header row (contains "GROUP" or datetime objects in same row)
    header_row = None
    for i, row in enumerate(rows):
        cells = [c for c in row if c is not None]
        strs  = [str(c).strip().upper() for c in cells]
        if "GROUP" in strs or any("GROUP" in s for s in strs):
            header_row = i
            break
    if header_row is None:
        return None

    import datetime
    date_labels_xlsx = []
    for val in rows[header_row]:
        if isinstance(val, datetime.datetime):
            date_labels_xlsx.append(val.strftime("%b %-d"))
        elif isinstance(val, str) and val.strip() and val.strip().upper() != "GROUP":
            date_labels_xlsx.append(val.strip())

    if not date_labels_xlsx:
        date_labels_xlsx = [f"Session {i+1}" for i in range(3)]

    groups: list[str] = []
    sessions: list[list[float]] = [[] for _ in date_labels_xlsx]

    for row in rows[header_row + 1:]:
        label = row[1] if len(row) > 1 else None
        if label is None:
            continue
        label_str = str(label).strip()
        if not label_str or label_str.upper() in ("TOTAL", "GRAND TOTAL", ""):
            continue
        if "TOTAL" in label_str.upper() or "CHILDREN" in label_str.upper():
            continue
        vals = [row[j] if j < len(row) else None for j in range(2, 2 + len(date_labels_xlsx))]
        numeric = [float(v) if isinstance(v, (int, float)) else 0.0 for v in vals]
        if all(v == 0 for v in numeric):
            continue
        groups.append(label_str.title())
        for idx, v in enumerate(numeric):
            sessions[idx].append(v)

    if not groups:
        return None

    x     = np.arange(len(groups))
    n_s   = len(date_labels_xlsx)
    width = min(0.8 / n_s, 0.25)
    colors = [C_NAVY, C_GOLD, C_SKY]

    fig, ax = make_church_fig(figsize=(14, 6))
    for si in range(n_s):
        offset = -width * (n_s - 1) / 2 + si * width
        bars = ax.bar(x + offset, sessions[si], width,
                      label=date_labels_xlsx[si] if si < len(date_labels_xlsx) else f"Session {si+1}",
                      color=colors[si % len(colors)], zorder=3)
        for bar, v in zip(bars, sessions[si]):
            if v > 0:
                ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.3,
                        str(int(v)), ha="center", va="bottom",
                        fontsize=7, fontweight="bold", color=C_NAVY)

    ax.set_xticks(x)
    ax.set_xticklabels(groups, rotation=35, ha="right", fontsize=8.5)
    church_style(fig, ax, "Sabbath School Attendance by Group")
    ax.set_ylabel("Attendance", fontsize=10)
    ax.legend(title="Session", fontsize=9, title_fontsize=9,
               framealpha=0.9, edgecolor=C_GREY)
    add_watermark(fig)
    return fig


# ── 2. Sabbath School compound totals trend ──────────────────────────────────

def plot_sabbath_school_totals_trend(df: pd.DataFrame) -> Optional[Figure]:
    """
    Line chart of Sabbath School morning/afternoon compound totals.
    Reads the same SABBATH SCHOOL DATA*.xlsx.  If unavailable, falls back
    to the weekly sabbath_school_attendance column trend.
    """
    path = _find_xlsx("sabbath school")
    if path is not None:
        try:
            return _sabbath_school_totals_from_xlsx(path)
        except Exception as exc:
            log(f"Note: could not parse SS totals from XLSX ({exc}), falling back.")

    if not has_data(df, ["sabbath_school_attendance"]):
        return None
    labels = date_labels(df)
    fig, ax = make_church_fig()
    ax.plot(labels, df["sabbath_school_attendance"], "o-",
            color=C_NAVY, linewidth=2.5, markersize=8, zorder=3)
    for xi, yi in zip(labels, df["sabbath_school_attendance"].fillna(0)):
        if yi > 0:
            ax.annotate(str(int(yi)), (xi, yi),
                        textcoords="offset points", xytext=(0, 10),
                        ha="center", fontsize=9, fontweight="bold", color=C_NAVY)
    church_style(fig, ax, "Sabbath School Attendance Totals Trend")
    ax.set_ylabel("Total Attendance", fontsize=10)
    ax.tick_params(axis="x", rotation=25)
    add_watermark(fig)
    return fig


def _sabbath_school_totals_from_xlsx(path: Path) -> Optional[Figure]:
    import datetime

    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    if ws is None:
        return None
    rows = list(ws.iter_rows(values_only=True))

    # Collect all date columns and their TOTAL IN CHURCH COMPOUND rows
    header_row = None
    for i, row in enumerate(rows):
        if any(str(c).strip().upper() == "GROUP" for c in row if c is not None):
            header_row = i
            break
    if header_row is None:
        return None

    date_cols: list[str] = []
    col_indices: list[int] = []
    for ci, val in enumerate(rows[header_row]):
        if isinstance(val, datetime.datetime):
            date_cols.append(val.strftime("%b %-d"))
            col_indices.append(ci)
        elif isinstance(val, str) and re.match(r"\d{1,2}/\d{1,2}/\d{4}", val.strip()):
            date_cols.append(val.strip())
            col_indices.append(ci)

    if not date_cols:
        return None

    morning_totals: list[float] = [0.0] * len(date_cols)
    afternoon_totals: list[float] = [0.0] * len(date_cols)
    in_afternoon = False

    for row in rows[header_row + 1:]:
        label = str(row[1]).strip().upper() if len(row) > 1 and row[1] else ""
        if "AFTERNOON" in label or "PM" in label:
            in_afternoon = True
        if "TOTAL IN CHURCH" in label or ("TOTAL" in label and "COMPOUND" in label):
            for si, ci in enumerate(col_indices):
                v = row[ci] if ci < len(row) else None
                val = float(v) if isinstance(v, (int, float)) else 0.0
                if in_afternoon:
                    afternoon_totals[si] = val
                else:
                    morning_totals[si] = val

    if all(v == 0 for v in morning_totals):
        return None

    fig, ax = make_church_fig(figsize=(8, 5))
    ax.plot(date_cols, morning_totals, "o-", color=C_NAVY, linewidth=2.5, markersize=8,
            label="Morning Session", zorder=3)
    if any(v > 0 for v in afternoon_totals):
        ax.plot(date_cols, afternoon_totals, "s--", color=C_GOLD, linewidth=2.5, markersize=8,
                label="Afternoon Session", zorder=3)
    for xi, yi in zip(date_cols, morning_totals):
        ax.annotate(str(int(yi)), (xi, yi),
                    textcoords="offset points", xytext=(0, 10),
                    ha="center", fontsize=10, fontweight="bold", color=C_NAVY)
    for xi, yi in zip(date_cols, afternoon_totals):
        if yi > 0:
            ax.annotate(str(int(yi)), (xi, yi),
                        textcoords="offset points", xytext=(0, -16),
                        ha="center", fontsize=10, fontweight="bold", color=C_GOLD)

    church_style(fig, ax, "Sabbath School – Church Compound Totals")
    ax.set_ylabel("Total People", fontsize=10)
    ax.legend(fontsize=9, framealpha=0.9, edgecolor=C_GREY)
    ax.set_ylim(0, max(morning_totals) * 1.25)
    add_watermark(fig)
    return fig


# ── 3. Home Church total attendance (horizontal bar, ranked) ─────────────────

def _load_home_church_xlsx() -> Optional[pd.DataFrame]:
    """
    Parse HOME CHURCH DATA*.xlsx.
    Returns a DataFrame with columns:
        home_church, adults_m, adults_f, youth_m, youth_f,
        amb_m, amb_f, children_m, children_f, visitors_m, visitors_f, total
    """
    path = _find_xlsx("home church")
    if path is None:
        return None
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        if ws is None:
            return None
        rows = list(ws.iter_rows(values_only=True))
    except Exception as exc:
        log(f"Note: could not open Home Church XLSX ({exc}).")
        return None

    records = []
    for row in rows[12:]:          # first 12 rows are headers / metadata
        name = row[1] if len(row) > 1 else None
        if name is None:
            continue
        name_str = str(name).strip()
        if not name_str or "TOTAL" in name_str.upper():
            continue
        vals = [row[j] if j < len(row) else None for j in range(2, 12)]
        def g(i): return float(vals[i]) if isinstance(vals[i], (int, float)) else 0.0
        records.append({
            "home_church":  name_str.title(),
            "adults_m":     g(0), "adults_f": g(1),
            "youth_m":      g(2), "youth_f":  g(3),
            "amb_m":        g(4), "amb_f":    g(5),
            "children_m":   g(6), "children_f": g(7),
            "visitors_m":   g(8), "visitors_f": g(9),
        })
    if not records:
        return None
    hc = pd.DataFrame(records)
    hc["total"] = hc[["adults_m", "adults_f", "youth_m", "youth_f",
                       "amb_m", "amb_f", "children_m", "children_f",
                       "visitors_m", "visitors_f"]].sum(axis=1)
    return hc


def plot_home_church_attendance(df: pd.DataFrame) -> Optional[Figure]:
    """Horizontal bar chart of Home Church total attendance, ranked."""
    hc = _load_home_church_xlsx()
    if hc is None or hc.empty:
        return None

    hc = hc.sort_values("total", ascending=True)
    colors = [C_NAVY if i % 2 == 0 else C_SKY for i in range(len(hc))]

    fig, ax = make_church_fig(figsize=(10, 7))
    bars = ax.barh(hc["home_church"], hc["total"], color=colors, height=0.65, zorder=3)
    for bar in bars:
        w = bar.get_width()
        ax.text(w + max(hc["total"]) * 0.01, bar.get_y() + bar.get_height() / 2,
                str(int(w)), va="center", ha="left",
                fontsize=9, fontweight="bold", color=C_NAVY)
    church_style(fig, ax, "Home Church Attendance — Q1 2026")
    ax.set_xlabel("Total Attendance", fontsize=10)
    ax.set_xlim(0, max(hc["total"]) * 1.12)
    ax.yaxis.grid(False)
    ax.xaxis.grid(True, color=C_GREY, linewidth=0.6, linestyle="--", alpha=0.8)
    add_watermark(fig)
    return fig


# ── 4. Home Church stacked breakdown ────────────────────────────────────────

def plot_home_church_stacked(df: pd.DataFrame) -> Optional[Figure]:
    """Stacked bar chart showing category breakdown per home church."""
    hc = _load_home_church_xlsx()
    if hc is None or hc.empty:
        return None

    hc = hc.sort_values("total", ascending=False)

    categories = {
        "Adults (M)":   hc["adults_m"].tolist(),
        "Adults (F)":   hc["adults_f"].tolist(),
        "Youth (M)":    hc["youth_m"].tolist(),
        "Youth (F)":    hc["youth_f"].tolist(),
        "Ambassadors":  (hc["amb_m"] + hc["amb_f"]).tolist(),
        "Children":     (hc["children_m"] + hc["children_f"]).tolist(),
        "Visitors":     (hc["visitors_m"] + hc["visitors_f"]).tolist(),
    }
    cat_colors = [C_NAVY, C_SKY, C_GREEN, "#88C070", C_GOLD, C_ROSE, C_SLATE]

    x = np.arange(len(hc))
    fig, ax = make_church_fig(figsize=(13, 6))
    bottoms = np.zeros(len(hc))
    for (label, vals), col in zip(categories.items(), cat_colors):
        ax.bar(x, vals, 0.65, bottom=bottoms, label=label, color=col, zorder=3)
        bottoms += np.array(vals)

    ax.set_xticks(x)
    ax.set_xticklabels(hc["home_church"], rotation=35, ha="right", fontsize=8)
    church_style(fig, ax, "Home Church Attendance — Category Breakdown")
    ax.set_ylabel("Attendance", fontsize=10)
    ax.legend(fontsize=8.5, ncol=4, loc="upper right", framealpha=0.9, edgecolor=C_GREY)
    add_watermark(fig)
    return fig


# ── 5. Business Meeting per home church ─────────────────────────────────────

def _load_business_meeting_xlsx() -> Optional[pd.DataFrame]:
    path = _find_xlsx("business attendance")
    if path is None:
        return None
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        if ws is None:
            return None
        rows = list(ws.iter_rows(values_only=True))
    except Exception as exc:
        log(f"Note: could not open Business Meeting XLSX ({exc}).")
        return None

    records = []
    for row in rows:
        if not isinstance(row[0], int):
            continue
        name = row[1] if len(row) > 1 else None
        if name is None:
            continue
        att = float(row[4]) if len(row) > 4 and isinstance(row[4], (int, float)) else 0.0
        exp = float(row[6]) if len(row) > 6 and isinstance(row[6], (int, float)) else 0.0
        if exp > 0 or att > 0:
            records.append({"home_church": str(name).title(), "attended": att, "expected": exp})
    return pd.DataFrame(records) if records else None


def plot_business_meeting_per_home_church(df: pd.DataFrame) -> Optional[Figure]:
    """Grouped bar: Business Meeting attended vs expected per home church."""
    bm = _load_business_meeting_xlsx()
    if bm is None or bm.empty:
        # Fallback: use weekly board_business_meeting columns
        return plot_board_business_meeting_expected_vs_attended(df)

    bm = bm.sort_values("expected", ascending=False)
    x = np.arange(len(bm))
    w = 0.35

    fig, ax = make_church_fig(figsize=(13, 6))
    ax.bar(x - w / 2, bm["expected"], w, label="Expected", color=C_GREY,  zorder=3)
    ax.bar(x + w / 2, bm["attended"], w, label="Attended",  color=C_NAVY, zorder=3)

    for xi, (att, exp) in enumerate(zip(bm["attended"], bm["expected"])):
        if exp > 0:
            pct = att / exp * 100
            ax.text(xi, max(att, exp) + ax.get_ylim()[1] * 0.01,
                    f"{pct:.0f}%", ha="center", va="bottom",
                    fontsize=8, fontweight="bold",
                    color=C_GREEN if pct >= 80 else C_ROSE)
        ax.text(xi + w / 2, att / 2 if att > 0 else 0,
                str(int(att)), ha="center", va="center",
                fontsize=8, fontweight="bold", color=C_WHITE)

    ax.set_xticks(x)
    ax.set_xticklabels(bm["home_church"], rotation=35, ha="right", fontsize=8)
    church_style(fig, ax, "Business Meeting Attendance vs Expected — Q1 2026")
    ax.set_ylabel("People", fontsize=10)
    ax.legend(fontsize=9, framealpha=0.9, edgecolor=C_GREY)
    add_watermark(fig)
    return fig


# ── 6. Active Leaders — Board Meeting monthly trend ──────────────────────────

def plot_active_leaders_board(df: pd.DataFrame) -> Optional[Figure]:
    """
    Monthly board meeting attended vs expected (56).
    Uses board_business_meeting_attendance from the weekly data, grouped by
    month, OR parses the BOARD MEETING ATTENDANCE*.xlsx directly.
    """
    path = _find_xlsx("board meeting")
    if path is not None:
        try:
            return _board_meeting_from_xlsx(path)
        except Exception as exc:
            log(f"Note: could not parse Board Meeting XLSX ({exc}), falling back to weekly data.")

    if not has_data(df, ["board_business_meeting_attendance"]):
        return None

    temp = df[["week_start_date", "board_business_meeting_attendance",
               "board_business_meeting_expected"]].copy()
    temp["month"] = pd.to_datetime(temp["week_start_date"], errors="coerce").dt.to_period("M")
    monthly = temp.groupby("month", sort=True).agg(
        attended=("board_business_meeting_attendance", "sum"),
        expected=("board_business_meeting_expected", "sum"),
    ).reset_index()
    if monthly.empty:
        return None

    months   = [str(m) for m in monthly["month"]]
    attended = monthly["attended"].tolist()
    expected = [e if e > 0 else 56 for e in monthly["expected"]]  # default board size
    pcts     = [a / e * 100 if e > 0 else 0 for a, e in zip(attended, expected)]
    return _draw_board_chart(months, attended, expected, pcts)


def _board_meeting_from_xlsx(path: Path) -> Optional[Figure]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    if ws is None:
        return None
    rows = list(ws.iter_rows(values_only=True))

    months, attended, expected = [], [], []
    MONTH_NAMES = {"january", "february", "march", "april", "may", "june",
                   "july", "august", "september", "october", "november", "december"}
    for row in rows:
        # Month name is in column index 1 (B)
        label = row[1] if len(row) > 1 else None
        if label is None:
            continue
        label_str = str(label).strip()
        if label_str.lower() not in MONTH_NAMES:
            continue
        att = float(row[2]) if len(row) > 2 and isinstance(row[2], (int, float)) else 0.0
        exp = float(row[4]) if len(row) > 4 and isinstance(row[4], (int, float)) else 56.0
        months.append(label_str.title())
        attended.append(att)
        expected.append(exp if exp > 0 else 56.0)

    if not months:
        return None

    pcts = [a / e * 100 if e > 0 else 0 for a, e in zip(attended, expected)]
    return _draw_board_chart(months, attended, expected, pcts)


def _draw_board_chart(months: list, attended: list, expected: list, pcts: list) -> Figure:
    x = np.arange(len(months))
    w = 0.32

    fig, ax = make_church_fig(figsize=(9, 5))
    ax.bar(x - w / 2, expected, w, label=f"Expected",  color=C_GREY,  zorder=3)
    ax.bar(x + w / 2, attended, w, label="Attended",   color=C_NAVY,  zorder=3)

    for xi, (att, exp, pct) in enumerate(zip(attended, expected, pcts)):
        ax.text(xi + w / 2, att + max(attended) * 0.02,
                f"{pct:.0f}%", ha="center", va="bottom",
                fontsize=10, fontweight="bold",
                color=C_GREEN if pct >= 80 else C_ROSE)
        if att > 0:
            ax.text(xi + w / 2, att / 2,
                    str(int(att)), ha="center", va="center",
                    fontsize=10, fontweight="bold", color=C_WHITE)

    # Trend line over attended
    ax.plot(x + w / 2, attended, "D--", color=C_ROSE, linewidth=1.8, markersize=0, zorder=5)

    ax.set_xticks(x)
    ax.set_xticklabels(months, fontsize=10)
    ax.set_ylim(0, max(expected) * 1.3)
    church_style(fig, ax, "Active Church Leaders — Board Meeting Attendance 2026")
    ax.set_ylabel("Number of Leaders", fontsize=10)
    ax.legend(fontsize=9.5, framealpha=0.9, edgecolor=C_GREY)
    add_watermark(fig)
    return fig


# ── 7. Church composition donut ──────────────────────────────────────────────

def plot_church_composition_donut(df: pd.DataFrame) -> Optional[Figure]:
    """Donut chart of average weekly attendance by category."""
    parts = [c for c in ["men", "women", "youth", "children"] if has_data(df, [c])]
    if not parts:
        return None

    avgs   = [df[c].mean() for c in parts]
    labels = [c.title() for c in parts]
    colors = [C_NAVY, C_ROSE, C_GREEN, C_GOLD][:len(parts)]
    total  = sum(avgs)

    fig, ax = plt.subplots(figsize=(7, 6))
    fig.patch.set_facecolor(C_WHITE)
    pie_result = ax.pie(
        avgs, labels=labels, autopct="%1.1f%%",
        colors=colors, startangle=90,
        wedgeprops=dict(width=0.45, edgecolor=C_WHITE, linewidth=2),
        pctdistance=0.75,
        textprops=dict(fontsize=11),
    )
    # ax.pie() returns (wedges, label_texts, autotext_texts) when autopct is set
    autotexts: list = pie_result[2] if len(pie_result) > 2 else []
    for t in autotexts:
        t.set_color(C_WHITE)
        t.set_fontweight("bold")
        t.set_fontsize(10)

    ax.text(0, 0, f"{int(total):,}\nAvg/Week",
            ha="center", va="center",
            fontsize=12, fontweight="bold", color=C_NAVY)
    ax.set_title("Average Weekly Church Composition — Q1 2026",
                 fontsize=13, fontweight="bold", color=C_NAVY, pad=16)
    add_watermark(fig)
    return fig


# ── 8. Weekly demographic trend (styled) ────────────────────────────────────

def plot_demographic_trends_styled(df: pd.DataFrame) -> Optional[Figure]:
    """Multi-line weekly attendance trend with church palette + fill."""
    columns = [c for c in ["men", "women", "youth", "children"] if has_data(df, [c])]
    if len(columns) < 2:
        return None

    labels  = date_labels(df)
    markers = ["o", "s", "^", "D"]
    colors  = [C_NAVY, C_ROSE, C_GREEN, C_GOLD]

    fig, ax = make_church_fig(figsize=(12, 5.5))
    for col, marker, color in zip(columns, markers, colors):
        ax.plot(labels, df[col], f"{marker}-",
                color=color, linewidth=2.5, markersize=7,
                label=col.title(), zorder=3)
        ax.fill_between(labels, df[col], alpha=0.06, color=color)

    church_style(fig, ax, "Weekly Sabbath Attendance by Category — Q1 2026")
    ax.set_ylabel("Attendance Count", fontsize=10)
    ax.set_xlabel("Sabbath Date", fontsize=10)
    ax.legend(fontsize=9.5, framealpha=0.9, edgecolor=C_GREY, loc="upper left")
    ax.tick_params(axis="x", rotation=25)
    add_watermark(fig)
    return fig


# ── 9. Financial trend (styled) ──────────────────────────────────────────────

def plot_financial_trend_styled(df: pd.DataFrame) -> Optional[Figure]:
    """Tithe & Offerings trend with fill under curves."""
    if not has_data(df, ["tithe", "offerings"]):
        return None

    labels = date_labels(df)
    fig, ax = make_church_fig(figsize=(12, 5))
    ax.fill_between(labels, df["tithe"].fillna(0),    alpha=0.18, color=C_NAVY)
    ax.fill_between(labels, df["offerings"].fillna(0), alpha=0.18, color=C_GOLD)
    ax.plot(labels, df["tithe"],    "o-",  color=C_NAVY, linewidth=2.5, markersize=7,
            label="Tithe", zorder=3)
    ax.plot(labels, df["offerings"], "s--", color=C_GOLD, linewidth=2.5, markersize=7,
            label="Offerings", zorder=3)
    church_style(fig, ax, "Weekly Tithe & Offerings — Q1 2026")
    ax.set_ylabel("KES", fontsize=10)
    ax.legend(fontsize=9.5, framealpha=0.9, edgecolor=C_GREY)
    ax.tick_params(axis="x", rotation=25)
    add_watermark(fig)
    return fig


# ── 10. Holy Communion per home church ───────────────────────────────────────

def _load_holy_communion_xlsx() -> Optional[pd.DataFrame]:
    path = _find_xlsx("holy communion")
    if path is None:
        return None
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        if ws is None:
            return None
        rows = list(ws.iter_rows(values_only=True))
    except Exception as exc:
        log(f"Note: could not open Holy Communion XLSX ({exc}).")
        return None

    records = []
    for row in rows:
        if not isinstance(row[0], int):
            continue
        name = row[1] if len(row) > 1 else None
        if name is None:
            continue
        att = float(row[2]) if len(row) > 2 and isinstance(row[2], (int, float)) else 0.0
        exp = float(row[3]) if len(row) > 3 and isinstance(row[3], (int, float)) else 0.0
        records.append({"home_church": str(name).title(), "attended": att, "expected": exp})
    return pd.DataFrame(records) if records else None


def plot_holy_communion_per_home_church(df: pd.DataFrame) -> Optional[Figure]:
    """Grouped bar: Holy Communion attended vs expected per home church."""
    hc = _load_holy_communion_xlsx()
    if hc is None or hc.empty:
        return plot_holy_communion_expected_vs_attended(df)

    hc = hc.sort_values("expected", ascending=False)
    x = np.arange(len(hc))
    w = 0.35

    fig, ax = make_church_fig(figsize=(14, 6))
    ax.bar(x - w / 2, hc["expected"], w, label="Expected", color=C_GREY, zorder=3)
    ax.bar(x + w / 2, hc["attended"], w, label="Attended",  color=C_GOLD, zorder=3)

    for xi, (att, exp) in enumerate(zip(hc["attended"], hc["expected"])):
        if exp > 0:
            pct = att / exp * 100
            ax.text(xi, max(att, exp) + ax.get_ylim()[1] * 0.01,
                    f"{pct:.0f}%", ha="center", va="bottom",
                    fontsize=7, fontweight="bold",
                    color=C_GREEN if pct >= 50 else C_ROSE)

    ax.set_xticks(x)
    ax.set_xticklabels(hc["home_church"], rotation=38, ha="right", fontsize=8)
    church_style(fig, ax, "Holy Communion Q1 2026 — Attendance vs Expected")
    ax.set_ylabel("People", fontsize=10)
    ax.legend(fontsize=9, framealpha=0.9, edgecolor=C_GREY)
    add_watermark(fig)
    return fig


# ---------------------------------------------------------------------------
# Graph builders
# ---------------------------------------------------------------------------


def plot_total_attendance_trend(df: pd.DataFrame) -> Optional[Figure]:
    if not has_data(df, ["total_attendance"]):
        return None
    labels = date_labels(df)
    fig, ax = plt.subplots(figsize=(11, 5))
    ax.plot(labels, df["total_attendance"], marker="o", linewidth=2.2, color="#2563eb")
    add_mean_line(ax, df["total_attendance"])
    configure_axes(ax, "Total Attendance Trend", "Attendance")
    ax.tick_params(axis="x", rotation=25)
    ax.legend()
    return fig


def plot_demographic_trends(df: pd.DataFrame) -> Optional[Figure]:
    columns = [c for c in ATTENDANCE_PARTS if has_data(df, [c])]
    if len(columns) < 2:
        return None
    labels = date_labels(df)
    fig, ax = plt.subplots(figsize=(11, 5))
    for column in columns:
        ax.plot(labels, df[column], marker="o", linewidth=1.8, label=column.replace("_", " ").title())
    configure_axes(ax, "Attendance Trends by Group", "Attendance")
    ax.tick_params(axis="x", rotation=25)
    ax.legend()
    return fig


def plot_attendance_by_group(df: pd.DataFrame) -> Optional[Figure]:
    columns = [c for c in ATTENDANCE_PARTS if has_data(df, [c])]
    if len(columns) < 2:
        return None
    labels = date_labels(df)
    x = np.arange(len(labels))
    width = min(0.8 / len(columns), 0.18)
    fig, ax = plt.subplots(figsize=(12, 5.5))
    offset_start = -width * (len(columns) - 1) / 2
    for index, column in enumerate(columns):
        ax.bar(x + offset_start + index * width, df[column].fillna(0), width, label=column.replace("_", " ").title())
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=25, ha="right")
    configure_axes(ax, "Attendance by Group per Week", "Attendance")
    ax.legend(ncol=min(len(columns), 3))
    return fig


def plot_attendance_distribution(df: pd.DataFrame) -> Optional[Figure]:
    columns = [c for c in ATTENDANCE_PARTS if has_data(df, [c])]
    if not columns:
        return None
    totals = [numeric_series(df, c).sum() for c in columns]
    if sum(totals) <= 0:
        return None
    fig, ax = plt.subplots(figsize=(7, 7))
    ax.pie(totals, labels=[c.replace("_", " ").title() for c in columns], autopct="%1.1f%%", startangle=90)
    ax.set_title("Attendance Distribution")
    return fig


def plot_attendance_growth(df: pd.DataFrame) -> Optional[Figure]:
    if not has_data(df, ["attendance_growth"]):
        return None
    labels = date_labels(df)
    values = df["attendance_growth"].fillna(0)
    colors = ["#16a34a" if value >= 0 else "#dc2626" for value in values]
    fig, ax = plt.subplots(figsize=(11, 5))
    ax.bar(labels, values, color=colors)
    ax.axhline(0, color="#111111", linewidth=0.8)
    configure_axes(ax, "Week-over-Week Attendance Growth", "Growth %")
    ax.tick_params(axis="x", rotation=25)
    return fig


def plot_adult_vs_young(df: pd.DataFrame) -> Optional[Figure]:
    if not has_data(df, ["adult_attendance", "young_attendance"]):
        return None
    labels = date_labels(df)
    x = np.arange(len(labels))
    width = 0.35
    fig, ax = plt.subplots(figsize=(11, 5))
    ax.bar(x - width / 2, df["adult_attendance"], width, label="Adults")
    ax.bar(x + width / 2, df["young_attendance"], width, label="Young")
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=25, ha="right")
    configure_axes(ax, "Adult vs Young Attendance", "Attendance")
    ax.legend()
    return fig


def plot_demographic_percentage_trends(df: pd.DataFrame) -> Optional[Figure]:
    columns = ["men_pct", "women_pct", "youth_pct", "children_pct"]
    if not has_data(df, columns):
        return None
    labels = date_labels(df)
    fig, ax = plt.subplots(figsize=(11, 5))
    for column in columns:
        ax.plot(labels, df[column], marker="o", linewidth=1.8, label=column.replace("_pct", "").title())
    configure_axes(ax, "Demographic Percentage Trends", "Share of Sabbath Attendance (%)")
    ax.tick_params(axis="x", rotation=25)
    ax.legend()
    return fig


def plot_average_demographic_percentage(df: pd.DataFrame) -> Optional[Figure]:
    columns = ["men_pct", "women_pct", "youth_pct", "children_pct"]
    if not has_data(df, columns):
        return None
    fig, ax = plt.subplots(figsize=(8, 5))
    ax.bar(["Men", "Women", "Youth", "Children"], [df[c].mean() for c in columns], color="#2563eb")
    configure_axes(ax, "Average Demographic Percentage", "Percentage (%)")
    return fig


def plot_sabbath_school_trend(df: pd.DataFrame) -> Optional[Figure]:
    if not has_data(df, ["sabbath_school_attendance"]):
        return None
    labels = date_labels(df)
    fig, ax = plt.subplots(figsize=(11, 5))
    ax.plot(labels, df["sabbath_school_attendance"], marker="o", linewidth=2, color="#7c3aed")
    add_mean_line(ax, df["sabbath_school_attendance"])
    configure_axes(ax, "Sabbath School Attendance Trend", "Attendance")
    ax.tick_params(axis="x", rotation=25)
    ax.legend()
    return fig


def plot_visitors_trend(df: pd.DataFrame) -> Optional[Figure]:
    if not has_data(df, ["visitors_count"]):
        return None
    labels = date_labels(df)
    fig, ax = plt.subplots(figsize=(11, 5))
    ax.bar(labels, df["visitors_count"].fillna(0), color="#0f766e")
    configure_axes(ax, "Visitors per Week", "Visitors")
    ax.tick_params(axis="x", rotation=25)
    return fig


def plot_tithe_offerings_trend(df: pd.DataFrame) -> Optional[Figure]:
    if not has_data(df, ["tithe", "offerings"]):
        return None
    labels = date_labels(df)
    fig, ax = plt.subplots(figsize=(11, 5))
    ax.plot(labels, df["tithe"], marker="o", linewidth=2, label="Tithe")
    ax.plot(labels, df["offerings"], marker="s", linewidth=2, label="Offerings")
    configure_axes(ax, "Tithe vs Offerings Trend", "Amount")
    ax.tick_params(axis="x", rotation=25)
    ax.legend()
    return fig


def plot_income_composition(df: pd.DataFrame) -> Optional[Figure]:
    columns = [c for c in INCOME_PARTS if has_data(df, [c])]
    if len(columns) < 2:
        return None
    labels = date_labels(df)
    fig, ax = plt.subplots(figsize=(11, 5))
    ax.stackplot(labels, [df[c].fillna(0) for c in columns], labels=[c.replace("_", " ").title() for c in columns], alpha=0.85)
    configure_axes(ax, "Income Composition over Time", "Amount")
    ax.tick_params(axis="x", rotation=25)
    ax.legend(loc="upper left")
    return fig


def plot_income_distribution(df: pd.DataFrame) -> Optional[Figure]:
    columns = [c for c in INCOME_PARTS if has_data(df, [c])]
    if not columns:
        return None
    totals = [numeric_series(df, c).sum() for c in columns]
    if sum(totals) <= 0:
        return None
    fig, ax = plt.subplots(figsize=(7, 7))
    ax.pie(totals, labels=[c.replace("_", " ").title() for c in columns], autopct="%1.1f%%", startangle=90)
    ax.set_title("Income Distribution")
    return fig


def plot_income_vs_attendance_dual(df: pd.DataFrame) -> Optional[Figure]:
    if not has_data(df, ["total_attendance", "total_income"]):
        return None
    labels = date_labels(df)
    fig, ax1 = plt.subplots(figsize=(11, 5))
    ax2 = ax1.twinx()
    ax1.plot(labels, df["total_attendance"], marker="o", color="#2563eb", label="Attendance")
    ax2.plot(labels, df["total_income"], marker="s", color="#16a34a", label="Income")
    ax1.set_title("Total Attendance vs Total Income")
    ax1.set_ylabel("Attendance")
    ax2.set_ylabel("Income")
    ax1.tick_params(axis="x", rotation=25)
    lines, labels_combined = [], []
    for axis in (ax1, ax2):
        handles, axis_labels = axis.get_legend_handles_labels()
        lines.extend(handles)
        labels_combined.extend(axis_labels)
    ax1.legend(lines, labels_combined, loc="upper left")
    ax1.grid(True, alpha=0.25)
    return fig


def pairwise_bar(df: pd.DataFrame, left: str, right: str, title: str, ylabel: str = "Amount") -> Optional[Figure]:
    if not has_data(df, [left, right]):
        return None
    labels = date_labels(df)
    x = np.arange(len(labels))
    width = 0.35
    fig, ax = plt.subplots(figsize=(11, 5))
    ax.bar(x - width / 2, df[left].fillna(0), width, label=left.replace("_", " ").title())
    ax.bar(x + width / 2, df[right].fillna(0), width, label=right.replace("_", " ").title())
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=25, ha="right")
    configure_axes(ax, title, ylabel)
    ax.legend()
    return fig


def plot_regular_vs_total_income(df: pd.DataFrame) -> Optional[Figure]:
    return pairwise_bar(df, "regular_income", "total_income", "Regular vs Total Income per Week")


def plot_income_growth(df: pd.DataFrame) -> Optional[Figure]:
    if not has_data(df, ["income_growth"]):
        return None
    labels = date_labels(df)
    values = df["income_growth"].fillna(0)
    colors = ["#16a34a" if value >= 0 else "#dc2626" for value in values]
    fig, ax = plt.subplots(figsize=(11, 5))
    ax.bar(labels, values, color=colors)
    ax.axhline(0, color="#111111", linewidth=0.8)
    configure_axes(ax, "Week-over-Week Income Growth", "Growth %")
    ax.tick_params(axis="x", rotation=25)
    return fig


def plot_per_capita_metrics(df: pd.DataFrame) -> Optional[Figure]:
    columns = ["income_per_attendee", "tithe_per_attendee", "offerings_per_attendee"]
    if not has_data(df, columns):
        return None
    labels = date_labels(df)
    fig, ax = plt.subplots(figsize=(11, 5))
    for column in columns:
        ax.plot(labels, df[column], marker="o", linewidth=1.8, label=column.replace("_", " ").title())
    configure_axes(ax, "Per Capita Giving Metrics", "Amount")
    ax.tick_params(axis="x", rotation=25)
    ax.legend()
    return fig


def plot_tithe_per_attendee(df: pd.DataFrame) -> Optional[Figure]:
    if not has_data(df, ["tithe_per_attendee"]):
        return None
    labels = date_labels(df)
    fig, ax = plt.subplots(figsize=(11, 5))
    ax.bar(labels, df["tithe_per_attendee"].fillna(0), color="#16a34a")
    configure_axes(ax, "Tithe per Attendee", "Amount")
    ax.tick_params(axis="x", rotation=25)
    return fig


def plot_regular_income_per_adult(df: pd.DataFrame) -> Optional[Figure]:
    if not has_data(df, ["regular_income_per_adult"]):
        return None
    labels = date_labels(df)
    fig, ax = plt.subplots(figsize=(11, 5))
    ax.bar(labels, df["regular_income_per_adult"].fillna(0), color="#2563eb")
    configure_axes(ax, "Regular Income per Adult", "Amount")
    ax.tick_params(axis="x", rotation=25)
    return fig


def plot_ratio_trends(df: pd.DataFrame) -> Optional[Figure]:
    columns = ["men_women_ratio", "adult_young_ratio", "tithe_offerings_ratio"]
    if not has_data(df, columns):
        return None
    labels = date_labels(df)
    fig, ax = plt.subplots(figsize=(11, 5))
    for column in columns:
        ax.plot(labels, df[column], marker="o", linewidth=1.8, label=column.replace("_", " ").title())
    configure_axes(ax, "Ratio Trends", "Ratio")
    ax.tick_params(axis="x", rotation=25)
    ax.legend()
    return fig


def plot_baptisms_trend(df: pd.DataFrame) -> Optional[Figure]:
    if not has_data(df, ["baptisms"]):
        return None
    labels = date_labels(df)
    fig, ax = plt.subplots(figsize=(11, 5))
    ax.bar(labels, df["baptisms"].fillna(0), color="#0284c7")
    configure_axes(ax, "Baptisms per Week", "Baptisms")
    ax.tick_params(axis="x", rotation=25)
    return fig


def plot_holy_communion_trend(df: pd.DataFrame) -> Optional[Figure]:
    if not has_data(df, ["holy_communion"]):
        return None
    labels = date_labels(df)
    fig, ax = plt.subplots(figsize=(11, 5))
    ax.plot(labels, df["holy_communion"], marker="o", linewidth=2, color="#9333ea")
    configure_axes(ax, "Holy Communion Attendance", "Attendance")
    ax.tick_params(axis="x", rotation=25)
    return fig


def plot_baptisms_vs_holy_communion(df: pd.DataFrame) -> Optional[Figure]:
    return pairwise_bar(df, "baptisms", "holy_communion", "Baptisms vs Holy Communion", "Count")


def plot_holy_communion_expected_vs_attended(df: pd.DataFrame) -> Optional[Figure]:
    return pairwise_bar(
        df,
        "holy_communion",
        "holy_communion_expected",
        "Holy Communion: Attended vs Expected",
        "Count",
    )


def plot_ambassadors_trend(df: pd.DataFrame) -> Optional[Figure]:
    if not has_data(df, ["ambassadors_attendance"]):
        return None
    labels = date_labels(df)
    fig, ax = plt.subplots(figsize=(11, 5))
    ax.plot(labels, df["ambassadors_attendance"], marker="o", linewidth=2, color="#f59e0b")
    add_mean_line(ax, df["ambassadors_attendance"])
    configure_axes(ax, "Ambassadors Attendance Trend", "Attendance")
    ax.tick_params(axis="x", rotation=25)
    ax.legend()
    return fig


def plot_board_business_meeting_trend(df: pd.DataFrame) -> Optional[Figure]:
    if not has_data(df, ["board_business_meeting_attendance"]):
        return None
    labels = date_labels(df)
    fig, ax = plt.subplots(figsize=(11, 5))
    ax.plot(labels, df["board_business_meeting_attendance"], marker="o", linewidth=2, color="#0ea5e9")
    add_mean_line(ax, df["board_business_meeting_attendance"])
    configure_axes(ax, "Board/Business Meeting Attendance", "Attendance")
    ax.tick_params(axis="x", rotation=25)
    ax.legend()
    return fig


def plot_board_business_meeting_expected_vs_attended(df: pd.DataFrame) -> Optional[Figure]:
    return pairwise_bar(
        df,
        "board_business_meeting_attendance",
        "board_business_meeting_expected",
        "Board/Business Meeting: Attended vs Expected",
        "Attendance",
    )


def plot_attendance_income_scatter(df: pd.DataFrame) -> Optional[Figure]:
    clean = clean_positive_pairs(df, "total_attendance", "total_income")
    if len(clean) < 2:
        return None
    fig, ax = plt.subplots(figsize=(8, 6))
    ax.scatter(clean["total_attendance"], clean["total_income"], color="#2563eb", s=70, alpha=0.75)
    if clean["total_attendance"].nunique() > 1:
        fit = np.polyfit(clean["total_attendance"], clean["total_income"], 1)
        line = np.poly1d(fit)
        xs = np.linspace(clean["total_attendance"].min(), clean["total_attendance"].max(), 100)
        ax.plot(xs, line(xs), "--", color="#dc2626", linewidth=1.5)
    configure_axes(ax, "Attendance vs Income Correlation", "Total Income")
    ax.set_xlabel("Total Attendance")
    return fig


def dual_axis(df: pd.DataFrame, primary: str, secondary: str, title: str) -> Optional[Figure]:
    if not has_data(df, [primary, secondary]):
        return None
    labels = date_labels(df)
    fig, ax1 = plt.subplots(figsize=(11, 5))
    ax2 = ax1.twinx()
    ax1.plot(labels, df[primary], marker="o", color="#2563eb", label=primary.replace("_", " ").title())
    ax2.plot(labels, df[secondary], marker="s", color="#f97316", label=secondary.replace("_", " ").title())
    ax1.set_title(title)
    ax1.set_ylabel(primary.replace("_", " ").title())
    ax2.set_ylabel(secondary.replace("_", " ").title())
    ax1.tick_params(axis="x", rotation=25)
    ax1.grid(True, alpha=0.25)
    lines, labels_combined = [], []
    for axis in (ax1, ax2):
        handles, axis_labels = axis.get_legend_handles_labels()
        lines.extend(handles)
        labels_combined.extend(axis_labels)
    ax1.legend(lines, labels_combined, loc="upper left")
    return fig


def plot_correlation_heatmap(df: pd.DataFrame) -> Optional[Figure]:
    columns = [
        "men",
        "women",
        "youth",
        "children",
        "sunday_home_church",
        "total_attendance",
        "tithe",
        "offerings",
        "total_income",
        "baptisms",
        "holy_communion",
        "sabbath_school_attendance",
    ]
    available = [column for column in columns if has_data(df, [column])]
    if len(available) < 3:
        return None
    corr = df[available].corr(numeric_only=True)
    fig, ax = plt.subplots(figsize=(10, 8))
    image = ax.imshow(corr, cmap="coolwarm", vmin=-1, vmax=1)
    ax.set_xticks(np.arange(len(available)))
    ax.set_yticks(np.arange(len(available)))
    ax.set_xticklabels([c.replace("_", " ").title() for c in available], rotation=45, ha="right")
    ax.set_yticklabels([c.replace("_", " ").title() for c in available])
    for row in range(len(available)):
        for col in range(len(available)):
            ax.text(col, row, f"{corr.iloc[row, col]:.2f}", ha="center", va="center", fontsize=8)
    ax.set_title("Metric Correlation Heatmap")
    fig.colorbar(image, ax=ax, shrink=0.8)
    return fig


def plot_distribution_histograms(df: pd.DataFrame) -> Optional[Figure]:
    columns = [
        "men",
        "women",
        "youth",
        "children",
        "sunday_home_church",
        "total_attendance",
        "tithe",
        "offerings",
        "total_income",
    ]
    available = [column for column in columns if has_data(df, [column])]
    if not available:
        return None
    rows = int(np.ceil(len(available) / 3))
    fig, axes = plt.subplots(rows, 3, figsize=(13, max(4, rows * 3.4)))
    axes_flat = np.atleast_1d(axes).ravel()
    for axis, column in zip(axes_flat, available):
        values = numeric_series(df, column).dropna()
        axis.hist(values, bins=min(8, max(3, len(values))), color="#2563eb", alpha=0.78)
        axis.axvline(values.mean(), color="#dc2626", linestyle="--", linewidth=1, label="Mean")
        axis.set_title(column.replace("_", " ").title())
        axis.grid(True, alpha=0.2)
    for axis in axes_flat[len(available) :]:
        axis.axis("off")
    fig.suptitle("Distribution Histograms", y=1.02)
    return fig


def plot_attendance_moving_average(df: pd.DataFrame) -> Optional[Figure]:
    if not has_data(df, ["total_attendance"]) or len(df) < 2:
        return None
    labels = date_labels(df)
    window = min(4, max(2, len(df) // 3 or 2))
    moving = df["total_attendance"].rolling(window=window, min_periods=1).mean()
    fig, ax = plt.subplots(figsize=(11, 5))
    ax.plot(labels, df["total_attendance"], marker="o", label="Actual")
    ax.plot(labels, moving, marker="s", linewidth=2.2, label=f"{window}-Week Moving Average")
    configure_axes(ax, "Attendance Moving Average", "Attendance")
    ax.tick_params(axis="x", rotation=25)
    ax.legend()
    return fig


def plot_attendance_forecast(df: pd.DataFrame) -> Optional[Figure]:
    clean = df[df["week_start_date"].notna() & df["total_attendance"].notna()].copy()
    if len(clean) < 2:
        return None
    clean = clean.sort_values("week_start_date")
    y = clean["total_attendance"].to_numpy(dtype=float)
    x = np.arange(len(clean), dtype=float)
    slope, intercept = np.polyfit(x, y, 1)
    future_steps = np.arange(len(clean), len(clean) + 4, dtype=float)
    future_dates = [clean["week_start_date"].iloc[-1] + pd.Timedelta(days=7 * (i + 1)) for i in range(4)]
    forecast = np.maximum(slope * future_steps + intercept, 0)

    fig, ax = plt.subplots(figsize=(11, 5))
    ax.plot(clean["week_start_date"], y, marker="o", linewidth=2, label="Historical")
    ax.plot(future_dates, forecast, marker="s", linestyle="--", linewidth=2, label="Forecast")
    configure_axes(ax, "Four-Week Attendance Forecast", "Attendance")
    fig.autofmt_xdate()
    ax.legend()
    return fig


def plot_summary_dashboard(df: pd.DataFrame) -> Optional[Figure]:
    if not has_data(df, ["total_attendance", "total_income"]):
        return None
    labels = date_labels(df)
    fig, axes = plt.subplots(2, 2, figsize=(13, 8))

    axes[0, 0].plot(labels, df["total_attendance"], marker="o", color="#2563eb")
    configure_axes(axes[0, 0], "Attendance", "Count")
    axes[0, 0].tick_params(axis="x", rotation=25)

    axes[0, 1].plot(labels, df["total_income"], marker="s", color="#16a34a")
    configure_axes(axes[0, 1], "Income", "Amount")
    axes[0, 1].tick_params(axis="x", rotation=25)

    attendance_cols = [c for c in ATTENDANCE_PARTS if has_data(df, [c])]
    attendance_totals = [df[c].sum() for c in attendance_cols]
    if sum(attendance_totals) > 0:
        axes[1, 0].pie(attendance_totals, labels=[c.replace("_", " ").title() for c in attendance_cols], autopct="%1.1f%%")
    axes[1, 0].set_title("Attendance Mix")

    income_cols = [c for c in INCOME_PARTS if has_data(df, [c])]
    income_totals = [df[c].sum() for c in income_cols]
    if sum(income_totals) > 0:
        axes[1, 1].pie(income_totals, labels=[c.replace("_", " ").title() for c in income_cols], autopct="%1.1f%%")
    axes[1, 1].set_title("Income Mix")

    fig.suptitle("Church Analytics Summary", fontsize=15)
    return fig


GraphBuilder = Callable[[pd.DataFrame], Optional[Figure]]


@dataclass(frozen=True)
class GraphSpec:
    graph_id: str
    title: str
    group: str
    filename: str
    builder: GraphBuilder
    description: str


GRAPH_SPECS: dict[str, GraphSpec] = {
    "total_attendance_trend": GraphSpec(
        "total_attendance_trend",
        "Total Attendance Trend",
        "attendance",
        "total_attendance_trend.png",
        plot_total_attendance_trend,
        "Line chart of app-style total attendance over time.",
    ),
    "demographic_trends": GraphSpec(
        "demographic_trends",
        "Attendance Trends by Group",
        "attendance",
        "demographic_trends.png",
        plot_demographic_trends,
        "Multi-line weekly trend for attendance groups.",
    ),
    "attendance_by_group": GraphSpec(
        "attendance_by_group",
        "Attendance by Group per Week",
        "attendance",
        "attendance_by_group.png",
        plot_attendance_by_group,
        "Grouped bar chart for Men/Women/Youth/Children/Home Church.",
    ),
    "attendance_distribution": GraphSpec(
        "attendance_distribution",
        "Attendance Distribution",
        "attendance",
        "attendance_distribution.png",
        plot_attendance_distribution,
        "Pie chart of aggregate attendance mix.",
    ),
    "attendance_growth": GraphSpec(
        "attendance_growth",
        "Attendance Growth",
        "attendance",
        "attendance_growth.png",
        plot_attendance_growth,
        "Week-over-week attendance growth percentage.",
    ),
    "adult_vs_young": GraphSpec(
        "adult_vs_young",
        "Adult vs Young Attendance",
        "attendance",
        "adult_vs_young.png",
        plot_adult_vs_young,
        "Grouped bar chart of adults against youth and children.",
    ),
    "demographic_percentage_trends": GraphSpec(
        "demographic_percentage_trends",
        "Demographic Percentage Trends",
        "attendance",
        "demographic_percentage_trends.png",
        plot_demographic_percentage_trends,
        "Weekly percentage share of Sabbath attendance groups.",
    ),
    "average_demographic_percentage": GraphSpec(
        "average_demographic_percentage",
        "Average Demographic Percentage",
        "attendance",
        "average_demographic_percentage.png",
        plot_average_demographic_percentage,
        "Average percentage share for Men/Women/Youth/Children.",
    ),
    "sabbath_school_trend": GraphSpec(
        "sabbath_school_trend",
        "Sabbath School Attendance Trend",
        "attendance",
        "sabbath_school_trend.png",
        plot_sabbath_school_trend,
        "Trend for optional Sabbath School attendance.",
    ),
    "visitors_trend": GraphSpec(
        "visitors_trend",
        "Visitors Trend",
        "attendance",
        "visitors_trend.png",
        plot_visitors_trend,
        "Weekly visitor counts when available.",
    ),
    "ambassadors_trend": GraphSpec(
        "ambassadors_trend",
        "Ambassadors Attendance Trend",
        "attendance",
        "ambassadors_trend.png",
        plot_ambassadors_trend,
        "Weekly ambassadors attendance when available.",
    ),
    "tithe_offerings_trend": GraphSpec(
        "tithe_offerings_trend",
        "Tithe vs Offerings Trend",
        "financial",
        "tithe_offerings_trend.png",
        plot_tithe_offerings_trend,
        "Line chart comparing tithe and offerings over time.",
    ),
    "income_composition": GraphSpec(
        "income_composition",
        "Income Composition",
        "financial",
        "income_composition.png",
        plot_income_composition,
        "Stacked area chart of giving streams.",
    ),
    "income_distribution": GraphSpec(
        "income_distribution",
        "Income Distribution",
        "financial",
        "income_distribution.png",
        plot_income_distribution,
        "Pie chart of aggregate income mix.",
    ),
    "income_vs_attendance": GraphSpec(
        "income_vs_attendance",
        "Income vs Attendance",
        "financial",
        "income_vs_attendance.png",
        plot_income_vs_attendance_dual,
        "Dual-axis trend for attendance and income.",
    ),
    "tithe_vs_offerings_week": GraphSpec(
        "tithe_vs_offerings_week",
        "Tithe vs Offerings per Week",
        "financial",
        "tithe_vs_offerings_week.png",
        lambda df: pairwise_bar(df, "tithe", "offerings", "Tithe vs Offerings per Week"),
        "Grouped bar comparison of tithe and offerings.",
    ),
    "regular_vs_total_income": GraphSpec(
        "regular_vs_total_income",
        "Regular vs Total Income",
        "financial",
        "regular_vs_total_income.png",
        plot_regular_vs_total_income,
        "Grouped bar comparison of regular income and total income.",
    ),
    "income_growth": GraphSpec(
        "income_growth",
        "Income Growth",
        "financial",
        "income_growth.png",
        plot_income_growth,
        "Week-over-week total income growth percentage.",
    ),
    "per_capita_metrics": GraphSpec(
        "per_capita_metrics",
        "Per Capita Metrics",
        "financial",
        "per_capita_metrics.png",
        plot_per_capita_metrics,
        "Income, tithe, and offerings per attendee.",
    ),
    "tithe_per_attendee": GraphSpec(
        "tithe_per_attendee",
        "Tithe per Attendee",
        "financial",
        "tithe_per_attendee.png",
        plot_tithe_per_attendee,
        "Bar chart of tithe divided by total attendance.",
    ),
    "regular_income_per_adult": GraphSpec(
        "regular_income_per_adult",
        "Regular Income per Adult",
        "financial",
        "regular_income_per_adult.png",
        plot_regular_income_per_adult,
        "Bar chart of tithe plus offerings per adult attendee.",
    ),
    "ratio_trends": GraphSpec(
        "ratio_trends",
        "Ratio Trends",
        "financial",
        "ratio_trends.png",
        plot_ratio_trends,
        "Men:Women, Adult:Young, and Tithe:Offerings ratios.",
    ),
    "baptisms_trend": GraphSpec(
        "baptisms_trend",
        "Baptisms Trend",
        "events",
        "baptisms_trend.png",
        plot_baptisms_trend,
        "Weekly baptism counts when available.",
    ),
    "holy_communion_trend": GraphSpec(
        "holy_communion_trend",
        "Holy Communion Trend",
        "events",
        "holy_communion_trend.png",
        plot_holy_communion_trend,
        "Holy Communion attendance when available.",
    ),
    "baptisms_vs_holy_communion": GraphSpec(
        "baptisms_vs_holy_communion",
        "Baptisms vs Holy Communion",
        "events",
        "baptisms_vs_holy_communion.png",
        plot_baptisms_vs_holy_communion,
        "Grouped weekly event-count comparison.",
    ),
    "holy_communion_expected_vs_attended": GraphSpec(
        "holy_communion_expected_vs_attended",
        "Holy Communion: Attended vs Expected",
        "events",
        "holy_communion_expected_vs_attended.png",
        plot_holy_communion_expected_vs_attended,
        "Grouped comparison of holy communion expected vs attended.",
    ),
    "board_business_meeting_trend": GraphSpec(
        "board_business_meeting_trend",
        "Board/Business Meeting Attendance",
        "meetings",
        "board_business_meeting_trend.png",
        plot_board_business_meeting_trend,
        "Weekly board/business meeting attendance when available.",
    ),
    "board_business_meeting_expected_vs_attended": GraphSpec(
        "board_business_meeting_expected_vs_attended",
        "Board/Business Meeting: Attended vs Expected",
        "meetings",
        "board_business_meeting_expected_vs_attended.png",
        plot_board_business_meeting_expected_vs_attended,
        "Grouped bar comparison of expected vs attended board/business meetings.",
    ),
    "attendance_income_scatter": GraphSpec(
        "attendance_income_scatter",
        "Attendance vs Income Scatter",
        "correlation",
        "attendance_income_scatter.png",
        plot_attendance_income_scatter,
        "Scatter plot with trendline.",
    ),
    "men_vs_tithe": GraphSpec(
        "men_vs_tithe",
        "Men vs Tithe",
        "correlation",
        "men_vs_tithe.png",
        lambda df: dual_axis(df, "men", "tithe", "Men vs Tithe over Time"),
        "Dual-axis chart comparing men attendance and tithe.",
    ),
    "women_vs_offerings": GraphSpec(
        "women_vs_offerings",
        "Women vs Offerings",
        "correlation",
        "women_vs_offerings.png",
        lambda df: dual_axis(df, "women", "offerings", "Women vs Offerings over Time"),
        "Dual-axis chart comparing women attendance and offerings.",
    ),
    "correlation_heatmap": GraphSpec(
        "correlation_heatmap",
        "Correlation Heatmap",
        "correlation",
        "correlation_heatmap.png",
        plot_correlation_heatmap,
        "Correlation matrix heatmap for available numeric metrics.",
    ),
    "distribution_histograms": GraphSpec(
        "distribution_histograms",
        "Distribution Histograms",
        "advanced",
        "distribution_histograms.png",
        plot_distribution_histograms,
        "Histogram grid for attendance and finance metrics.",
    ),
    "attendance_moving_average": GraphSpec(
        "attendance_moving_average",
        "Attendance Moving Average",
        "advanced",
        "attendance_moving_average.png",
        plot_attendance_moving_average,
        "Actual attendance with a rolling average.",
    ),
    "attendance_forecast": GraphSpec(
        "attendance_forecast",
        "Attendance Forecast",
        "advanced",
        "attendance_forecast.png",
        plot_attendance_forecast,
        "Simple four-week linear attendance forecast.",
    ),
    "summary_dashboard": GraphSpec(
        "summary_dashboard",
        "Summary Dashboard",
        "advanced",
        "summary_dashboard.png",
        plot_summary_dashboard,
        "Compact 2x2 dashboard of attendance, income, and mix charts.",
    ),
    # ── Presentation-grade graphs (church palette) ────────────────────────
    "sabbath_school_groups": GraphSpec(
        "sabbath_school_groups",
        "Sabbath School Attendance by Group",
        "presentation",
        "sabbath_school_groups.png",
        plot_sabbath_school_groups,
        "Grouped bar for each SS group × session; reads SABBATH SCHOOL DATA*.xlsx.",
    ),
    "sabbath_school_totals_trend": GraphSpec(
        "sabbath_school_totals_trend",
        "Sabbath School Compound Totals Trend",
        "presentation",
        "sabbath_school_totals_trend.png",
        plot_sabbath_school_totals_trend,
        "Morning/afternoon compound totals trend from SABBATH SCHOOL DATA*.xlsx.",
    ),
    "home_church_attendance": GraphSpec(
        "home_church_attendance",
        "Home Church Total Attendance",
        "presentation",
        "home_church_attendance.png",
        plot_home_church_attendance,
        "Horizontal ranked bar for each home church total; reads HOME CHURCH DATA*.xlsx.",
    ),
    "home_church_stacked": GraphSpec(
        "home_church_stacked",
        "Home Church Category Breakdown",
        "presentation",
        "home_church_stacked.png",
        plot_home_church_stacked,
        "Stacked bar per home church (Adults/Youth/Ambassadors/Children/Visitors).",
    ),
    "business_meeting_per_home_church": GraphSpec(
        "business_meeting_per_home_church",
        "Business Meeting Attendance vs Expected",
        "presentation",
        "business_meeting_per_home_church.png",
        plot_business_meeting_per_home_church,
        "Grouped bar: business meeting attended vs expected per home church.",
    ),
    "active_leaders_board": GraphSpec(
        "active_leaders_board",
        "Active Leaders — Board Meeting",
        "presentation",
        "active_leaders_board.png",
        plot_active_leaders_board,
        "Monthly board meeting attended vs expected; reads BOARD MEETING*.xlsx.",
    ),
    "church_composition_donut": GraphSpec(
        "church_composition_donut",
        "Church Composition Donut",
        "presentation",
        "church_composition_donut.png",
        plot_church_composition_donut,
        "Donut chart of average weekly Men/Women/Youth/Children split.",
    ),
    "demographic_trends_styled": GraphSpec(
        "demographic_trends_styled",
        "Weekly Demographic Trends (Styled)",
        "presentation",
        "demographic_trends_styled.png",
        plot_demographic_trends_styled,
        "Multi-line weekly attendance trend with church palette and fill.",
    ),
    "financial_trend_styled": GraphSpec(
        "financial_trend_styled",
        "Weekly Financial Trend (Styled)",
        "presentation",
        "financial_trend_styled.png",
        plot_financial_trend_styled,
        "Tithe & offerings trend with fill under curves, church palette.",
    ),
    "holy_communion_per_home_church": GraphSpec(
        "holy_communion_per_home_church",
        "Holy Communion Attendance vs Expected",
        "presentation",
        "holy_communion_per_home_church.png",
        plot_holy_communion_per_home_church,
        "Grouped bar: holy communion attended vs expected per home church.",
    ),
}

GRAPH_GROUPS: dict[str, list[str]] = {
    group: [graph_id for graph_id, spec in GRAPH_SPECS.items() if spec.group == group]
    for group in sorted({spec.group for spec in GRAPH_SPECS.values()})
}
GRAPH_GROUPS["all"] = list(GRAPH_SPECS.keys())


def print_graph_list() -> None:
    log("Available graph groups:")
    for group in sorted(GRAPH_GROUPS):
        log(f"  {group}: {len(GRAPH_GROUPS[group])} graph(s)")
    log("\nAvailable graphs:")
    for graph_id, spec in GRAPH_SPECS.items():
        log(f"  {graph_id:<34} [{spec.group}] {spec.description}")


def parse_graph_selection(graph_values: Optional[Sequence[str]], group: Optional[str]) -> list[str]:
    selected: list[str] = []

    if group:
        selected.extend(GRAPH_GROUPS.get(group, []))

    if graph_values:
        for value in graph_values:
            for token in [part.strip().lower() for part in value.split(",") if part.strip()]:
                if token in GRAPH_GROUPS:
                    selected.extend(GRAPH_GROUPS[token])
                elif token in GRAPH_SPECS:
                    selected.append(token)
                else:
                    log(f"Warning: unknown graph id or group ignored: {token}")

    if not selected:
        selected = list(GRAPH_SPECS.keys())

    deduped: list[str] = []
    seen: set[str] = set()
    for graph_id in selected:
        if graph_id not in seen:
            deduped.append(graph_id)
            seen.add(graph_id)
    return deduped


def prompt_text(label: str, default: Optional[str] = None) -> Optional[str]:
    suffix = f" [{default}]" if default else ""
    value = input(f"{label}{suffix}: ").strip()
    return value or default


def prompt_bool(label: str, default: bool = False) -> bool:
    suffix = "Y/n" if default else "y/N"
    while True:
        value = input(f"{label} ({suffix}): ").strip().lower()
        if not value:
            return default
        if value in {"y", "yes"}:
            return True
        if value in {"n", "no"}:
            return False
        log("Please enter y or n.")


def prompt_int(label: str, default: Optional[int] = None, min_value: Optional[int] = None) -> Optional[int]:
    suffix = f" [{default}]" if default is not None else ""
    while True:
        value = input(f"{label}{suffix}: ").strip()
        if not value:
            return default
        try:
            parsed = int(value)
        except ValueError:
            log("Please enter a whole number.")
            continue
        if min_value is not None and parsed < min_value:
            log(f"Please enter a value >= {min_value}.")
            continue
        return parsed


def parse_csv_list(value: Optional[str]) -> list[str]:
    if not value:
        return []
    return [part.strip() for part in value.split(",") if part.strip()]


def apply_interactive_overrides(args: argparse.Namespace) -> argparse.Namespace:
    default_input_list = default_inputs()
    default_input_text = ", ".join(default_input_list) if default_input_list else None

    inputs: list[str] = []
    while not inputs:
        input_value = prompt_text("Input file(s) or folder(s) (comma-separated)", default_input_text)
        inputs = parse_csv_list(input_value) if input_value else []
        if not inputs:
            inputs = default_input_list
        if not inputs:
            log("Please provide at least one input path.")

    output_dir = prompt_text("Output directory", args.output_dir) or args.output_dir

    log("Available graph groups: " + ", ".join(sorted(GRAPH_GROUPS)))
    group: Optional[str] = None
    while True:
        group_value = prompt_text("Select graph group (leave blank to choose graph IDs)", None)
        if not group_value:
            break
        group_value = group_value.strip().lower()
        if group_value in GRAPH_GROUPS:
            group = group_value
            break
        log("Unknown group. Enter one of: " + ", ".join(sorted(GRAPH_GROUPS)))

    graphs: Optional[list[str]] = None
    if group is None:
        graph_value = prompt_text("Graph IDs (comma-separated, blank = all)", None)
        graphs = parse_csv_list(graph_value)
        if not graphs:
            graphs = None

    pdf = prompt_bool("Export PDF", args.pdf)
    show = prompt_bool("Show charts interactively", args.show)
    export_clean = prompt_bool("Export normalized CSV", args.export_clean)
    no_tables = prompt_bool("Skip summary tables", args.no_tables)
    dpi = prompt_int("PNG DPI", args.dpi, min_value=50) or args.dpi
    force_year = prompt_int("Force year (blank for none)", args.force_year, min_value=1900)

    args.input = inputs
    args.output_dir = output_dir
    args.group = group
    args.graphs = graphs
    args.pdf = pdf
    args.show = show
    args.export_clean = export_clean
    args.no_tables = no_tables
    args.dpi = dpi
    args.force_year = force_year
    return args


def save_stats_tables(df: pd.DataFrame, out_dir: Path) -> None:
    tables_dir = out_dir / "tables"
    tables_dir.mkdir(parents=True, exist_ok=True)

    numeric_cols = [
        column
        for column in df.columns
        if pd.api.types.is_numeric_dtype(df[column]) and df[column].notna().any()
    ]
    if numeric_cols:
        df[numeric_cols].describe().to_csv(tables_dir / "summary_statistics.csv")

    corr_cols = [
        "men",
        "women",
        "youth",
        "children",
        "sunday_home_church",
        "total_attendance",
        "tithe",
        "offerings",
        "total_income",
        "baptisms",
        "holy_communion",
        "sabbath_school_attendance",
    ]
    corr_available = [column for column in corr_cols if column in df.columns and df[column].notna().sum() >= 2]
    if len(corr_available) >= 2:
        df[corr_available].corr(numeric_only=True).to_csv(tables_dir / "correlation_matrix.csv")


def save_figure(fig: Figure, spec: GraphSpec, out_dir: Path, pdf: Optional[PdfPages], dpi: int) -> Path:
    group_dir = out_dir / spec.group
    group_dir.mkdir(parents=True, exist_ok=True)
    fig.tight_layout()
    path = group_dir / spec.filename
    fig.savefig(str(path), dpi=dpi, bbox_inches="tight")
    if pdf is not None:
        pdf.savefig(fig)
    return path


def generate_graphs(
    df: pd.DataFrame,
    graph_ids: Sequence[str],
    out_dir: Path,
    make_pdf: bool,
    show: bool,
    dpi: int,
) -> tuple[int, int]:
    pdf: Optional[PdfPages] = PdfPages(str(out_dir / "graphs.pdf")) if make_pdf else None
    generated = 0
    skipped = 0

    for graph_id in graph_ids:
        spec = GRAPH_SPECS[graph_id]
        try:
            fig = spec.builder(df)
        except Exception as exc:
            skipped += 1
            log(f"Skipping {graph_id}: graph failed ({exc})")
            continue

        if fig is None:
            skipped += 1
            log(f"Skipping {graph_id}: required data not available.")
            continue

        path = save_figure(fig, spec, out_dir, pdf, dpi)
        generated += 1
        log(f"Generated {path}")

        if show:
            if matplotlib.get_backend().lower() == "agg":
                log("Warning: interactive --show is unavailable on the current headless backend.")
            else:
                plt.show()
        plt.close(fig)

    if pdf is not None:
        pdf.close()
        log(f"Generated {out_dir / 'graphs.pdf'}")

    return generated, skipped


def default_inputs() -> list[str]:
    for candidate in ["data/netFinalData.csv", "data/finalData.csv", "data/church_data.csv"]:
        path = SCRIPT_DIR / candidate
        if path.exists():
            return [str(path)]
    return []


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Generate terminal/PDF church analytics graphs from CSV or XLSX data.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument(
        "--input",
        nargs="*",
        default=None,
        help="CSV/XLSX files or directories containing them.",
    )
    parser.add_argument(
        "--output-dir",
        "--output",
        default="church_analysis",
        help="Directory for normalized data, tables, PNG charts, and optional PDF.",
    )
    parser.add_argument(
        "--graphs",
        nargs="*",
        default=None,
        help="Graph IDs or comma-separated graph IDs/groups. Use 'all' for everything.",
    )
    parser.add_argument(
        "--group",
        choices=sorted(GRAPH_GROUPS),
        default=None,
        help="Generate one graph group.",
    )
    parser.add_argument("--list-graphs", action="store_true", help="List available graph IDs and exit.")
    parser.add_argument(
        "--interactive",
        action="store_true",
        help="Run in interactive mode to choose inputs, outputs, and graphs.",
    )
    parser.add_argument("--pdf", action="store_true", help="Export generated graphs into one PDF file.")
    parser.add_argument("--show", action="store_true", help="Display charts interactively when a GUI backend is available.")
    parser.add_argument("--dpi", type=int, default=150, help="PNG export DPI.")
    parser.add_argument("--force-year", type=int, default=None, help="Force parsed dates to this year.")
    parser.add_argument("--export-clean", action="store_true", help="Also write normalized weekly-record CSV.")
    parser.add_argument(
        "--no-tables",
        action="store_true",
        help="Skip summary statistics and correlation CSV tables.",
    )
    return parser


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = build_parser()
    argv = list(sys.argv[1:] if argv is None else argv)
    args = parser.parse_args(argv)

    if args.list_graphs:
        print_graph_list()
        return 0

    if args.interactive:
        args = apply_interactive_overrides(args)
    elif not argv and sys.stdin.isatty():
        if prompt_bool("Run in interactive mode?", False):
            args = apply_interactive_overrides(args)

    input_values = args.input if args.input else default_inputs()
    if not input_values:
        parser.error("No input files provided and no default data files found.")

    paths = resolve_input_paths(input_values)
    if not paths:
        parser.error("No readable CSV/XLSX files found in the provided inputs.")

    log(f"Loading {len(paths)} input file(s)...")
    df = load_data(paths, args.force_year)
    if df.empty:
        parser.error("No usable weekly records loaded. Check the input columns and data.")

    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    normalized_path = out_dir / "normalized_data.csv"
    if args.export_clean:
        export_cols = [column for column in APP_WEEKLY_COLUMNS if column in df.columns]
        export_cols += [
            column
            for column in OPTIONAL_METADATA_COLUMNS
            if column in df.columns and column not in export_cols
        ]
        extra_cols = [
            "sabbath_attendance",
            "total_with_home_church",
            "core_income",
            "regular_income",
            "special_collections",
            "income_per_attendee",
            "attendance_growth",
            "income_growth",
        ]
        ordered_cols = export_cols + [column for column in extra_cols if column in df.columns and column not in export_cols]
        df[ordered_cols].to_csv(normalized_path, index=False)
        log(f"Generated {normalized_path}")

    if not args.no_tables:
        save_stats_tables(df, out_dir)
        log(f"Generated summary tables in {out_dir / 'tables'}")

    graph_ids = parse_graph_selection(args.graphs, args.group)
    generated, skipped = generate_graphs(df, graph_ids, out_dir, args.pdf, args.show, args.dpi)

    log("")
    log(f"Loaded rows: {len(df)}")
    log(f"Generated graphs: {generated}")
    if skipped:
        log(f"Skipped graphs: {skipped}")
    log(f"Output directory: {out_dir}")
    return 0 if generated > 0 else 1


if __name__ == "__main__":
    plt.style.use("seaborn-v0_8-whitegrid")
    raise SystemExit(main())